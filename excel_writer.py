# Arquivo: core/excel_writer.py (VERSÃO SIMPLES)
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from core.logger import SistemaLog

class ExcelReportWriter:
    @staticmethod
    def gerar_relatorio(caminho_arquivo, df_concilia, df_nfe, df_cte):
        """
        Gera o arquivo Excel final. (Adaptado para ignorar conciliação se for None)
        """
        try:
            with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
                has_data = False
                
                # 1. Aba Resumo (SÓ CRIA SE TIVER DADOS E NÃO FOR NONE)
                if df_concilia is not None and not df_concilia.empty:
                    # Ordena por Status para erros ficarem no topo
                    if "Status Geral" in df_concilia.columns:
                        df_concilia.sort_values(by="Status Geral", ascending=True, inplace=True)
                    
                    df_concilia.to_excel(writer, sheet_name="RESUMO CONCILIAÇÃO", index=False)
                    has_data = True
                
                # 2. Aba NFe
                if not df_nfe.empty: 
                    df_nfe.to_excel(writer, sheet_name="NFe - Detalhado", index=False)
                    has_data = True

                # 3. Aba CTe
                if not df_cte.empty: 
                    df_cte.to_excel(writer, sheet_name="CTe - Detalhado", index=False)
                    has_data = True
                
                # Fallback se tudo vazio
                if not has_data:
                    pd.DataFrame(["Sem dados XML encontrados"]).to_excel(writer, sheet_name="Aviso", index=False)

                # Aplica estilos em todas as abas criadas
                for sheet in writer.sheets: 
                    ExcelReportWriter._estilizar_planilha(writer.sheets[sheet])
            
            return True

        except Exception as e:
            SistemaLog.registrar_erro("Erro ao gerar Excel formatado", e)
            raise e

    @staticmethod
    def _estilizar_planilha(ws):
        """Aplica as cores e formatações padrão"""
        
        # --- Definição de Cores e Estilos ---
        header_font = Font(name='Calibri', bold=True, color='FFFFFF')
        fill_header = PatternFill(start_color="00061A", end_color="00061A", fill_type='solid') # Azul Brasif
        
        # Cores para Status
        fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type='solid') # Verde
        fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type='solid')   # Vermelho
        fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type='solid') # Amarelo
        fill_gray = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type='solid')   # Cinza

        # 1. Congelar Painéis e Filtros
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        # 2. Formatar Cabeçalho
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 3. Ajuste de Colunas e Formatação Numérica
        for col in ws.columns:
            col_let = get_column_letter(col[0].column)
            head = str(col[0].value).upper()
            
            width = 15
            target_fmt = 'General'
            
            # A. Moeda
            palavras_moeda = [
                'VALOR', 'BASE', 'CREDITO', 'TOTAL', 'SAP', 'DIFERENÇA', 
                'FRETE', 'SEG', 'DESC', 'OUTR', 'PRECO', 'DESP'
            ]
            
            if any(x in head for x in palavras_moeda):
                width = 18
                target_fmt = '#,##0.00_-'
            
            # B. Porcentagem
            elif any(x in head for x in ['ALIQ', 'MVA', 'RED']): 
                width = 12
                target_fmt = '0.00%'
            
            # C. Larguras Específicas
            elif 'CHAVE' in head: width = 47
            elif 'STATUS' in head or 'CHECK' in head: width = 20
            elif any(x in head for x in ['PRODUTO', 'DESTINATARIO', 'TOMADOR', 'FORNECEDOR', 'EMITENTE', 'NATUREZA', 'OBSERVAÇÕES']): 
                width = 35
            elif any(x in head for x in ['CNPJ', 'EMP.', 'FIL.', 'UF', 'NCM', 'CEST', 'ANP', 'NUMERO', 'SERIE', 'MODELO']): 
                width = 12
            
            # Aplica a largura
            ws.column_dimensions[col_let].width = width
            
            # Aplica formato numérico nas células
            if target_fmt != 'General':
                for cell in col[1:]: 
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = target_fmt
            
            # 4. Formatação Condicional (Cores)
            if "EMP. XML" in head or "FIL. XML" in head or "ORIGEM" in head:
                for cell in col[1:]: cell.fill = fill_gray

            if "STATUS" in head or "CHECK" in head:
                for cell in col[1:]:
                    val = str(cell.value).upper()
                    if "OK" in val: cell.fill = fill_green
                    elif "DIVERGÊNCIA" in val or "ERRO" in val: cell.fill = fill_red
                    elif "SÓ NO" in val: cell.fill = fill_yellow
